import openpyxl
import pyautogui
import time
import random
import cv2
import pyperclip
import tkinter as tk
import os
from tkinter import messagebox

planilha = openpyxl.load_workbook('contatos.xlsx')
nav_aba = planilha.active


add = None
msg = None
numeros_nao_encontrados = []

def set_times():
    caixa = tk.Tk()
    caixa.title('BotWhatsApp')

    tk.Label(caixa, text='Tempo mínimo (segundos):').grid(row=0, column=0)
    entrada_min = tk.Entry(caixa)
    entrada_min.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(caixa, text='Tempo máximo (segundos):').grid(row=1, column=0)
    entrada_max = tk.Entry(caixa)
    entrada_max.grid(row=1, column=1, padx=5, pady=5)

    def abrir_planilha():
        nome_arquivo = 'contatos.xlsx'
        caminho_arquivo = os.path.abspath(nome_arquivo)
        os.startfile(caminho_arquivo)
        caixa.destroy()

    def ok():
        try:
            global tempo_min, tempo_max, add
            tempo_min = float(entrada_min.get())
            tempo_max = float(entrada_max.get())
            add_imagem_1 = pyautogui.locateOnScreen(
                'adicionar.png', confidence=0.7)
            add_imagem_2 = pyautogui.locateOnScreen(
                'adicionar2.png', confidence=0.7)
            if add_imagem_1 is not None:
                add = add_imagem_1
            elif add_imagem_2 is not None:
                add = add_imagem_2
            caixa.destroy()
        except Exception as e:
            if isinstance(e, ValueError):
                messagebox.showerror(
                    'Erro de validação', 'Tempo de envio não definido corretamente\n          (DIGITE APENAS NUMEROS)')

    tk.Button(caixa, text='EXECUTAR', command=ok, padx=7, pady=7).grid(
        row=2, column=0, padx=5, pady=5)

    tk.Button(caixa, text='ARQUIVO XLSX', command=abrir_planilha,
              padx=7, pady=7, foreground='#090').grid(row=2, column=1, padx=5, pady=5,)

    caixa.mainloop()


set_times()

if tempo_min and tempo_max and add:

    for linha in nav_aba.iter_rows(min_row=1, max_row=nav_aba.max_row, values_only=True):
        numero, mensagem = linha[:2]

        pyautogui.moveTo(add, duration=1)
        pyautogui.click()
        pyautogui.write(str(numero))
        pyautogui.hotkey('enter')
        time.sleep(1.5)
        msg_imagem_1 = pyautogui.locateOnScreen('mensagem.png', confidence=0.7)
        msg_imagem_2 = pyautogui.locateOnScreen(
            'mensagem2.png', confidence=0.7)
        if msg_imagem_1 is not None:
            msg = msg_imagem_1
        else:
            msg = msg_imagem_2

        aleatorio = random.uniform(tempo_min, tempo_max)
        row = list(nav_aba.iter_rows(values_only=True)).index(linha) + 1

        if msg:
            pyautogui.moveTo(pyautogui.center(msg), duration=1)
            pyautogui.click()
        else:
            numeros_nao_encontrados.append(numero)
            time.sleep(1)
            pyautogui.hotkey('esc')
            pyautogui.hotkey('esc')
            continue

        pyperclip.copy(str(mensagem))
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(aleatorio)
        pyautogui.hotkey('enter')
        time.sleep(0.5)
        pyautogui.hotkey('esc')

    def valores_lista():
        if numeros_nao_encontrados:
            mensagem = 'Numeros sem registro de WhatsApp:\n' + \
                '\n'.join(str(whatsapp)
                          for whatsapp in numeros_nao_encontrados)
        else:
            mensagem = 'Todos os numeros estão registrados no WhatsApp.'

        messagebox.showinfo("Lista de numeros", mensagem)
else:
    messagebox.showerror('Error:', 'Icone para chamar clientes não encontrado')

numeros_lista_excel = openpyxl.Workbook()

def criar_linhas(*args):
    args = list(args)
    sheet = numeros_lista_excel.active
    for numero in args:
        sheet.append([numero])
    numeros_lista_excel.save('caio_buono.xlsx')

criar_linhas(*numeros_nao_encontrados)
numeros_lista_excel.close()


valores_lista()
planilha.close()
